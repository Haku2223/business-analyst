"""GET /export — Export pipeline to Excel."""

import io
import logging
from datetime import datetime

from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import select

from app.auth import check_auth_redirect, get_display_name
from app.database import get_db
from app.models import Company, Note

logger = logging.getLogger(__name__)

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


@router.get("/export", response_class=HTMLResponse)
async def export_page(request: Request):
    """Export page."""
    if not check_auth_redirect(request):
        return RedirectResponse("/login", status_code=302)

    display_name = get_display_name(request)
    return templates.TemplateResponse(
        "export.html",
        {"request": request, "display_name": display_name, "active_page": "export"},
    )


@router.get("/export/download")
async def export_download(request: Request):
    """Stream an Excel file of the current pipeline."""
    if not check_auth_redirect(request):
        return RedirectResponse("/login", status_code=302)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return HTMLResponse("openpyxl not installed", status_code=500)

    async for db in get_db():
        result = await db.execute(select(Company))
        companies = result.scalars().all()

        # Organise by pipeline stage
        stages = {
            "Unreviewed": [c for c in companies if c.pipeline_status == "unreviewed"],
            "Watch": [c for c in companies if c.pipeline_status == "watch"],
            "Deep Dive": [c for c in companies if c.pipeline_status == "deep_dive"],
            "Pass": [c for c in companies if c.pipeline_status == "pass"],
        }

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        headers = [
            "ORG.NR", "Company Name", "City", "County", "Revenue (SEK)",
            "Employees", "Net Result (SEK)", "Profit Margin %", "Soliditet %",
            "Board Chair", "CEO", "Website", "Allabolag URL",
            "Pipeline Status", "Phase 2 Status", "AI Description",
        ]

        stage_colors = {
            "Unreviewed": "F3F4F6",
            "Watch": "DBEAFE",
            "Deep Dive": "DCFCE7",
            "Pass": "FEE2E2",
        }

        for stage, stage_companies in stages.items():
            ws = wb.create_sheet(title=stage)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=stage_colors.get(stage, "F3F4F6"))

            for c in stage_companies:
                ws.append([
                    c.orgnr,
                    c.bolagsnamn or "",
                    c.ort or "",
                    c.lan or "",
                    int(c.omsattning / 100) if c.omsattning else None,
                    c.antal_anstallda,
                    int(c.arets_resultat / 100) if c.arets_resultat else None,
                    c.vinstmarginal,
                    c.soliditet,
                    c.ordforande or "",
                    c.vd or "",
                    c.hemsida or "",
                    c.allabolag_url or "",
                    c.pipeline_status or "",
                    c.phase2_status or "",
                    c.ai_description or "",
                ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
        filename = f"pipeline_export_{ts}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
